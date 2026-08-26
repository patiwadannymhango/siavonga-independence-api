import io

import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from rest_framework import filters, status
from rest_framework.generics import ListAPIView, ListCreateAPIView, RetrieveUpdateAPIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend

from apps.common.permissions import IsAdminRole, IsStaffRole

from .models import VendorCategory, VendorRegistration
from .serializers import (
    AdminManualVendorRegistrationSerializer,
    AdminVendorCategorySerializer,
    AdminVendorRegistrationSerializer,
    AdminVendorRegistrationStatusUpdateSerializer,
    PublicVendorRegistrationCreateSerializer,
    VendorCategorySerializer,
    VendorRegistrationRecordSerializer,
)
from .services import create_vendor_registration

# ---------------------------------------------------------------------------
# Public
# ---------------------------------------------------------------------------


class PublicVendorCategoryListView(ListAPIView):
    """GET /api/v1/vendors/categories/"""

    permission_classes = [AllowAny]
    serializer_class = VendorCategorySerializer
    queryset = VendorCategory.objects.filter(is_active=True)
    pagination_class = None


class PublicVendorRegistrationCreateView(APIView):
    """POST /api/v1/vendors/register/"""

    permission_classes = [AllowAny]

    def post(self, request):
        serializer = PublicVendorRegistrationCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        vendor_registration = create_vendor_registration(**serializer.to_registration_kwargs())

        return Response(
            {
                "registrationId": vendor_registration.id,
                "reference": vendor_registration.registration_number,
                "amount": float(vendor_registration.amount),
                "currency": vendor_registration.currency,
                "status": vendor_registration.status,
            },
            status=status.HTTP_201_CREATED,
        )


class PublicVendorRegistrationLookupView(APIView):
    """GET /api/v1/vendors/lookup/?q=<reference-or-email>"""

    permission_classes = [AllowAny]

    def get(self, request):
        query = request.query_params.get("q", "").strip()

        if not query:
            return Response(
                {"detail": "Provide a reference number or email as ?q="},
                status=status.HTTP_400_BAD_REQUEST,
            )

        vendor_registration = (
            VendorRegistration.objects.select_related("vendor", "category")
            .filter(registration_number__iexact=query)
            .first()
            or VendorRegistration.objects.select_related("vendor", "category")
            .filter(vendor__email__iexact=query)
            .order_by("-registered_at")
            .first()
        )

        if not vendor_registration:
            return Response({"detail": "No matching registration found."}, status=status.HTTP_404_NOT_FOUND)

        return Response(VendorRegistrationRecordSerializer(vendor_registration).data)


# ---------------------------------------------------------------------------
# Admin-facing
# ---------------------------------------------------------------------------


class AdminVendorCategoryListView(ListCreateAPIView):
    """GET (any admin) / POST (ADMIN only) /api/v1/vendors/admin/categories/"""

    serializer_class = AdminVendorCategorySerializer
    queryset = VendorCategory.objects.all()
    pagination_class = None

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsStaffRole()]


class AdminVendorCategoryDetailView(RetrieveUpdateAPIView):
    """GET (any admin) / PATCH (ADMIN only) /api/v1/vendors/admin/categories/<id>/"""

    serializer_class = AdminVendorCategorySerializer
    queryset = VendorCategory.objects.all()

    def get_permissions(self):
        if self.request.method in ("PATCH", "PUT"):
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsStaffRole()]


class AdminVendorRegistrationListView(ListAPIView):
    """GET /api/v1/vendors/admin/registrations/"""

    permission_classes = [IsAuthenticated, IsStaffRole]
    serializer_class = AdminVendorRegistrationSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["status", "category"]
    search_fields = [
        "registration_number",
        "vendor__business_name",
        "vendor__full_name",
        "vendor__email",
        "vendor__phone",
    ]
    ordering_fields = ["registered_at", "amount", "status"]
    queryset = VendorRegistration.objects.select_related("vendor", "category")


class AdminVendorRegistrationDetailView(RetrieveUpdateAPIView):
    """
    GET   /api/v1/vendors/admin/registrations/<id>/ — any admin
    PATCH — ADMIN only. Accepts {"status": "..."}.
    """

    serializer_class = AdminVendorRegistrationSerializer
    queryset = VendorRegistration.objects.select_related("vendor", "category")

    def get_permissions(self):
        if self.request.method in ("PATCH", "PUT"):
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsStaffRole()]

    def patch(self, request, *args, **kwargs):
        vendor_registration = self.get_object()

        serializer = AdminVendorRegistrationStatusUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        new_status = serializer.validated_data["status"]
        old_status = vendor_registration.status

        vendor_registration.status = new_status
        vendor_registration.save(update_fields=["status", "updated_at"])

        if new_status == VendorRegistration.Status.CONFIRMED and old_status != new_status:
            vendor_registration.notify_confirmed()

        return Response(AdminVendorRegistrationSerializer(vendor_registration).data)


class AdminVendorRegistrationCreateView(APIView):
    """POST /api/v1/vendors/admin/registrations/manual/ — ADMIN only."""

    permission_classes = [IsAuthenticated, IsAdminRole]

    def post(self, request):
        serializer = AdminManualVendorRegistrationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        vendor_registration = create_vendor_registration(**serializer.to_registration_kwargs())

        return Response(AdminVendorRegistrationSerializer(vendor_registration).data, status=status.HTTP_201_CREATED)


class AdminVendorRegistrationExportView(APIView):
    """GET /api/v1/vendors/admin/registrations/export/"""

    permission_classes = [IsAuthenticated, IsStaffRole]

    COLUMNS = [
        ("Reference", lambda r: r.registration_number),
        ("Status", lambda r: r.status),
        ("Business name", lambda r: r.vendor.business_name),
        ("Contact person", lambda r: r.vendor.full_name),
        ("Email", lambda r: r.vendor.email),
        ("Phone", lambda r: r.vendor.phone),
        ("Business location", lambda r: r.vendor.business_location),
        ("Category", lambda r: r.category.name),
        ("Requirement", lambda r: r.requirement),
        ("Products/Services", lambda r: r.products_services),
        ("Amount", lambda r: float(r.amount)),
        ("Currency", lambda r: r.currency),
        ("Registered at", lambda r: r.registered_at.replace(tzinfo=None) if r.registered_at else None),
    ]

    def get(self, request):
        vendor_registrations = VendorRegistration.objects.select_related("vendor", "category").order_by(
            "-registered_at"
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Vendor Registrations"

        for col_index, (header, _) in enumerate(self.COLUMNS, start=1):
            sheet.cell(row=1, column=col_index, value=header)

        for row_index, vendor_registration in enumerate(vendor_registrations, start=2):
            for col_index, (_, getter) in enumerate(self.COLUMNS, start=1):
                sheet.cell(row=row_index, column=col_index, value=getter(vendor_registration))

        for col_index in range(1, len(self.COLUMNS) + 1):
            sheet.column_dimensions[get_column_letter(col_index)].width = 22

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="siavonga-run-vendor-registrations.xlsx"'
        return response
