import csv
import io

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl.utils import get_column_letter
from rest_framework import filters, status
from rest_framework.generics import ListAPIView, ListCreateAPIView, RetrieveUpdateAPIView, RetrieveUpdateDestroyAPIView
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.common.permissions import IsAdminRole, IsStaffRole

from .models import RaceCategory, Registration
from .serializers import (
    AdminManualRegistrationSerializer,
    AdminRaceCategorySerializer,
    AdminRegistrationSerializer,
    AdminRegistrationUpdateSerializer,
    PublicRegistrationCreateSerializer,
    RaceCategorySerializer,
    RegistrationRecordSerializer,
)
from .services import create_registration

# ---------------------------------------------------------------------------
# Public
# ---------------------------------------------------------------------------


class PublicRaceCategoryListView(ListAPIView):
    """GET /api/v1/categories/"""

    permission_classes = [AllowAny]
    serializer_class = RaceCategorySerializer
    queryset = RaceCategory.objects.filter(is_active=True)
    pagination_class = None


class PublicRegistrationCreateView(APIView):
    """POST /api/v1/registrations/"""

    permission_classes = [AllowAny]

    def post(self, request):
        serializer = PublicRegistrationCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        registration = create_registration(**serializer.to_registration_kwargs())

        return Response(
            {
                "registrationId": registration.id,
                "reference": registration.registration_number,
                # A JSON number, not a DRF-style decimal string — the
                # frontend calls .toFixed() on this directly.
                "amount": float(registration.amount),
                "currency": registration.currency,
            },
            status=status.HTTP_201_CREATED,
        )


class PublicRegistrationLookupView(APIView):
    """
    GET /api/v1/registrations/lookup/?q=<reference-or-email>

    Powers the "track your registration" feature. Deliberately returns
    only the fields the frontend's RegistrationRecord shape needs.
    """

    permission_classes = [AllowAny]

    def get(self, request):
        query = request.query_params.get("q", "").strip()

        if not query:
            return Response(
                {"detail": "Provide a reference number or email as ?q="},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registration = (
            Registration.objects.select_related("participant", "category")
            .filter(registration_number__iexact=query)
            .first()
            or Registration.objects.select_related("participant", "category")
            .filter(participant__email__iexact=query)
            .order_by("-registered_at")
            .first()
        )

        if not registration:
            return Response({"detail": "No matching registration found."}, status=status.HTTP_404_NOT_FOUND)

        return Response(RegistrationRecordSerializer(registration).data)


# ---------------------------------------------------------------------------
# Admin-facing
# ---------------------------------------------------------------------------


class AdminRaceCategoryListView(ListCreateAPIView):
    """
    GET  /api/v1/admin/categories/ — any admin (ADMIN or VIEW). Unlike the
    public endpoint, this includes inactive categories.

    POST — ADMIN only. Create a new race category.
    """

    serializer_class = AdminRaceCategorySerializer
    queryset = RaceCategory.objects.all()
    pagination_class = None

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsStaffRole()]


class AdminRaceCategoryDetailView(RetrieveUpdateAPIView):
    """
    GET   /api/v1/admin/categories/<id>/ — any admin.
    PATCH — ADMIN only. Edit a category's name, code, price, capacity, etc.,
    or flip `is_active` to retire it from the public site.

    No DELETE: Registration.category is on_delete=PROTECT, so a category
    with existing registrations can't be hard-deleted anyway — deactivate
    it instead.
    """

    serializer_class = AdminRaceCategorySerializer
    queryset = RaceCategory.objects.all()

    def get_permissions(self):
        if self.request.method in ("PATCH", "PUT"):
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsStaffRole()]


class AdminRegistrationListView(ListAPIView):
    """
    GET /api/v1/admin/registrations/

    Supports ?status=, ?category=, ?search=, ?ordering= and standard
    pagination — the data source for the admin registrations table.
    """

    permission_classes = [IsAuthenticated, IsStaffRole]
    serializer_class = AdminRegistrationSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["status", "category"]
    search_fields = [
        "registration_number",
        "participant__full_name",
        "participant__email",
        "participant__phone",
    ]
    ordering_fields = ["registered_at", "amount", "status"]
    queryset = Registration.objects.select_related("participant", "category").prefetch_related("payments")


class AdminRegistrationDetailView(RetrieveUpdateDestroyAPIView):
    """
    GET    /api/v1/admin/registrations/<id>/ — any admin account (ADMIN or read-only VIEW)

    PATCH  /api/v1/admin/registrations/<id>/ — ADMIN only. Partial update —
    any mix of `status` and participant/race detail fields (full_name,
    email, phone, gender, age_range, country, t_shirt_size,
    club_or_institution, emergency_contact_name, emergency_contact_phone,
    medical_notes). Flipping `status` to CONFIRMED (e.g. cash paid at the
    door, or a reconciled bank transfer) notifies the runner just like an
    online payment would, but only on the transition into confirmed.

    DELETE /api/v1/admin/registrations/<id>/ — ADMIN only. Removes the
    registration record (and any payments against it).
    """

    serializer_class = AdminRegistrationSerializer
    queryset = Registration.objects.select_related("participant", "category")

    def get_permissions(self):
        if self.request.method in ("PATCH", "PUT", "DELETE"):
            return [IsAuthenticated(), IsAdminRole()]
        return [IsAuthenticated(), IsStaffRole()]

    def perform_destroy(self, instance):
        # Payment.registration is on_delete=PROTECT, so clear those first
        # or Django raises ProtectedError instead of deleting.
        instance.payments.all().delete()
        instance.delete()

    def patch(self, request, *args, **kwargs):
        registration = self.get_object()

        serializer = AdminRegistrationUpdateSerializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        participant_updates = {k: v for k, v in data.items() if k in AdminRegistrationUpdateSerializer.PARTICIPANT_FIELDS}
        registration_updates = {k: v for k, v in data.items() if k not in AdminRegistrationUpdateSerializer.PARTICIPANT_FIELDS}

        if participant_updates:
            for field, value in participant_updates.items():
                setattr(registration.participant, field, value)
            registration.participant.save(update_fields=[*participant_updates.keys()])

        old_status = registration.status
        new_status = registration_updates.get("status")

        if registration_updates:
            for field, value in registration_updates.items():
                setattr(registration, field, value)
            registration.save(update_fields=[*registration_updates.keys(), "updated_at"])

        # Manually marking something CONFIRMED (e.g. cash paid at the
        # door, or a reconciled bank transfer) notifies the runner just
        # like an online payment would, but only on the transition into
        # confirmed.
        if new_status == Registration.Status.CONFIRMED and old_status != new_status:
            from apps.notifications.services import notify_payment_confirmed

            notify_payment_confirmed(registration)

        return Response(AdminRegistrationSerializer(registration).data)


class AdminRegistrationCreateView(APIView):
    """
    POST /api/v1/admin/registrations/manual/

    Manual "register a participant" — used by the admin for walk-ins,
    phone registrations, etc. ADMIN only.
    """

    permission_classes = [IsAuthenticated, IsAdminRole]

    def post(self, request):
        serializer = AdminManualRegistrationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        registration = create_registration(**serializer.to_registration_kwargs())

        return Response(AdminRegistrationSerializer(registration).data, status=status.HTTP_201_CREATED)


class AdminDashboardView(APIView):
    """
    GET /api/v1/admin/dashboard/

    One-stop summary for an admin dashboard: registration counts by
    status, revenue collected vs pending.
    """

    permission_classes = [IsAuthenticated, IsStaffRole]

    def get(self, request):
        from decimal import Decimal

        from django.db.models import Count, Sum

        registrations = Registration.objects.all()

        by_status = list(registrations.values("status").annotate(count=Count("id")))
        by_category = list(
            registrations.values("category__name").annotate(count=Count("id")).order_by("category__name")
        )

        today_count = registrations.filter(registered_at__date=timezone.localdate()).count()

        revenue_confirmed = (
            registrations.filter(status=Registration.Status.CONFIRMED).aggregate(total=Sum("amount"))["total"]
            or Decimal("0.00")
        )
        revenue_pending = (
            registrations.filter(
                status__in=[Registration.Status.PENDING_PAYMENT, Registration.Status.PAYMENT_PROCESSING]
            ).aggregate(total=Sum("amount"))["total"]
            or Decimal("0.00")
        )

        return Response(
            {
                "total_registrations": registrations.count(),
                "today_count": today_count,
                "by_status": by_status,
                "by_category": by_category,
                "revenue_confirmed": str(revenue_confirmed),
                "revenue_pending": str(revenue_pending),
            }
        )


class AdminRegistrationBulkUploadView(APIView):
    """
    POST /api/v1/admin/registrations/bulk-upload/

    Accepts a CSV or XLSX file (multipart field name "file") with
    columns: full_name, email, phone, category_code, status (optional,
    defaults to CONFIRMED). Returns a report of created rows and any rows
    that failed, rather than failing the whole batch on one bad row.
    ADMIN only.
    """

    permission_classes = [IsAuthenticated, IsAdminRole]
    parser_classes = [MultiPartParser]

    REQUIRED_COLUMNS = ["full_name", "category_code"]

    def post(self, request):
        upload = request.FILES.get("file")

        if not upload:
            return Response({"detail": "Attach a file under the 'file' field."}, status=status.HTTP_400_BAD_REQUEST)

        rows = self._parse_rows(upload)

        categories = {c.code: c for c in RaceCategory.objects.all()}

        created = []
        errors = []

        for index, row in enumerate(rows, start=2):  # header is row 1
            missing = [c for c in self.REQUIRED_COLUMNS if not row.get(c)]
            if missing:
                errors.append({"row": index, "error": f"Missing: {', '.join(missing)}"})
                continue

            category = categories.get(row["category_code"])
            if not category:
                errors.append({"row": index, "error": f"Unknown category_code '{row['category_code']}'"})
                continue

            try:
                registration = create_registration(
                    category=category,
                    participant_data={
                        "full_name": row["full_name"],
                        "email": row.get("email", ""),
                        "phone": row.get("phone", ""),
                    },
                    details={},
                )

                desired_status = (row.get("status") or "CONFIRMED").upper()
                if desired_status in Registration.Status.values and desired_status != registration.status:
                    registration.status = desired_status
                    registration.save(update_fields=["status", "updated_at"])

                created.append(registration.registration_number)

            except Exception as exc:  # noqa: BLE001
                errors.append({"row": index, "error": str(exc)})

        return Response(
            {
                "created_count": len(created),
                "created_references": created,
                "error_count": len(errors),
                "errors": errors,
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST,
        )

    def _parse_rows(self, upload):
        filename = (upload.name or "").lower()

        if filename.endswith(".csv"):
            text = upload.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            return [{(k or "").strip().lower(): (v or "").strip() for k, v in row.items()} for row in reader]

        workbook = openpyxl.load_workbook(upload, data_only=True)
        sheet = workbook.active

        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(h or "").strip().lower() for h in next(rows_iter)]

        rows = []
        for values in rows_iter:
            if all(v in (None, "") for v in values):
                continue
            row = {headers[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(values) if i < len(headers)}
            rows.append(row)

        return rows


class AdminRegistrationExportView(APIView):
    """
    GET /api/v1/admin/registrations/export/

    Streams an .xlsx of every registration.
    """

    permission_classes = [IsAuthenticated, IsStaffRole]

    COLUMNS = [
        ("Reference", lambda r: r.registration_number),
        ("Status", lambda r: r.status),
        ("Full name", lambda r: r.participant.full_name),
        ("Email", lambda r: r.participant.email),
        ("Phone", lambda r: r.participant.phone),
        ("Gender", lambda r: r.participant.gender),
        ("Age range", lambda r: r.participant.age_range),
        ("Category", lambda r: r.category.name),
        ("T-shirt size", lambda r: r.t_shirt_size),
        ("Club/Institution", lambda r: r.club_or_institution),
        ("Emergency contact", lambda r: r.emergency_contact_name),
        ("Emergency phone", lambda r: r.emergency_contact_phone),
        ("Amount", lambda r: float(r.amount)),
        ("Currency", lambda r: r.currency),
        ("Registered at", lambda r: r.registered_at.replace(tzinfo=None) if r.registered_at else None),
    ]

    def get(self, request):
        registrations = Registration.objects.select_related("participant", "category").order_by("-registered_at")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Registrations"

        for col_index, (header, _) in enumerate(self.COLUMNS, start=1):
            sheet.cell(row=1, column=col_index, value=header)

        for row_index, registration in enumerate(registrations, start=2):
            for col_index, (_, getter) in enumerate(self.COLUMNS, start=1):
                sheet.cell(row=row_index, column=col_index, value=getter(registration))

        for col_index in range(1, len(self.COLUMNS) + 1):
            sheet.column_dimensions[get_column_letter(col_index)].width = 22

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="siavonga-run-registrations.xlsx"'
        return response
